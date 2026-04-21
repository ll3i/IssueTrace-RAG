from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

import openpyxl


K_DOC_ID = "\uc758\uacb0\uc11c\uad00\ub9ac\ubc88\ud638"
K_TITLE = "\uc758\uacb0\uc11c\uc81c\ubaa9"
K_DATE = "\uacf5\uac1c\uc77c\uc790"
K_FILE_ID = "\uc758\uacb0\uc11c\ud30c\uc77c\uba85"
K_INFOS = "\ud53c\uc2ec\uc778\uc815\ubcf4"
K_MAIN = "\uc704\ubc18\uc720\ud615"
K_SUB = "\uc138\ubd80\uc704\ubc18\uc720\ud615"
K_ACTION = "\uc870\uce58\uc720\ud615"
K_ACTION_DATE = "\uc870\uce58\uc77c\uc790"
K_COMPANY = "\ud53c\uc2ec\uc778\uae30\uc5c5\uba85"

PERSONA_GENERAL = "\uc77c\ubc18\uad6d\ubbfc"
PERSONA_COMPLIANCE = "\uae30\uc5c5 \ucef4\ud50c\ub77c\uc774\uc5b8\uc2a4 \ub2f4\ub2f9\uc790"
PERSONA_STUDENT = "\ub17c\ubb38 \uc900\ube44 \ud559\uc0dd"

QUESTION_TYPES = [
    "\uc81c\uc7ac\uc911\uc2ec",
    "\ubc95\ub839\uc911\uc2ec",
    "\uc720\uc0ac\uc0ac\ub840",
    "\ud1b5\uacc4\uae30\ubc18",
    "\uc2dc\uc7a5\uc601\ud5a5",
]

DIFFICULTIES = ["\uae30\ucd08", "\uc911\uae09", "\uc2ec\ud654"]


LAW_MAP = {
    "\ubd88\uacf5\uc815\ud558\ub3c4\uae09\uac70\ub798\ud589\uc704": "\ud558\ub3c4\uae09\uac70\ub798 \uacf5\uc815\ud654\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\uc804\uc790\uc0c1\uac70\ub798\uc18c\ube44\uc790\ubcf4\ud638\ubc95\ub839 \uc704\ubc18": "\uc804\uc790\uc0c1\uac70\ub798 \ub4f1\uc5d0\uc11c\uc758 \uc18c\ube44\uc790\ubcf4\ud638\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\ubd80\ub2f9\ud55c \ud45c\uc2dc\uad11\uace0": "\ud45c\uc2dc\u00b7\uad11\uace0\uc758 \uacf5\uc815\ud654\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\ubd80\ub2f9\ud55c \uacf5\ub3d9\ud589\uc704": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\ub2e8\uccb4-\uacbd\uc7c1\uc81c\ud55c\ud589\uc704": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\uc0ac\uc5c5\uc790\ub2e8\uccb4\uae08\uc9c0\ud589\uc704": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\uac70\ub798\uc0c1\uc9c0\uc704\ub0a8\uc6a9": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\uac70\ub798\uc0c1 \uc9c0\uc704\uc758 \ub0a8\uc6a9": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\ubd88\uacf5\uc815\uac70\ub798\ud589\uc704": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\uad6c\uc18d\uc870\uac74\ubd80\uac70\ub798": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\ubd80\ub2f9\ud55c \uace0\uac1d\uc720\uc778": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\ubd80\ub2f9\ud55c \uc9c0\uc6d0": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\uac70\ub798\uac70\uc808": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\uc7ac\ud310\ub9e4\uac00\uaca9\uc720\uc9c0\ud589\uc704": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\uacbd\uc81c\ub825 \uc9d1\uc911\uc5b5\uc81c": "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\uac00\ub9f9\uc0ac\uc5c5\uac70\ub798\uc758\uacf5\uc815\ud654\uc5d0\uad00\ud55c\ubc95\ub839 \uc704\ubc18\uad00\ub828": "\uac00\ub9f9\uc0ac\uc5c5\uac70\ub798\uc758 \uacf5\uc815\ud654\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\ubc29\ubb38\ud310\ub9e4\ub4f1\uc5d0\uad00\ud55c\ubc95\ub839 \uc704\ubc18\uad00\ub828": "\ubc29\ubb38\ud310\ub9e4 \ub4f1\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\ub300\uaddc\ubaa8 \uc720\ud1b5\uc5c5\uc5d0\uc11c\uc758 \uac70\ub798 \uacf5\uc815\ud654\uc5d0 \uad00\ud55c \ubc95\ub960 \uc704\ubc18": "\ub300\uaddc\ubaa8\uc720\ud1b5\uc5c5\uc5d0\uc11c\uc758 \uac70\ub798 \uacf5\uc815\ud654\uc5d0 \uad00\ud55c \ubc95\ub960",
    "\ub300\ub9ac\uc810\uac70\ub798\uc758 \uacf5\uc815\ud654\uc5d0 \uad00\ud55c \ubc95\ub960 \uc704\ubc18": "\ub300\ub9ac\uc810\uac70\ub798\uc758 \uacf5\uc815\ud654\uc5d0 \uad00\ud55c \ubc95\ub960",
}


def uniq(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def normalize_text(text: str) -> str:
    text = text.replace("\\n", " ").replace("\n", " ")
    text = re.sub(r"\| --- .*", " ", text)
    text = re.sub(r"\|", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"^\-+\s*", "", text)
    return text


def shorten(text: str, limit: int = 140) -> str:
    text = normalize_text(text)
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "\u2026"


def clean_subject(text: str) -> str:
    text = re.sub(r"\[[^\]]+\]", "", text or "").strip()
    text = text.replace(".", "\u00b7")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ,")


def stable_variant(*parts: str, modulo: int) -> int:
    seed = "|".join(parts)
    return sum(ord(ch) for ch in seed) % modulo


def title_topic(title: str) -> str:
    topic = title
    topic = re.sub(r"\uc5d0 \ub300\ud55c \uac74\s*$", "", topic)
    if "\uad00\ub828" in topic:
        topic = topic.split("\uad00\ub828", 1)[-1].strip()
    elif "\uc758 " in topic:
        topic = topic.split("\uc758 ", 1)[-1].strip()
    return clean_subject(topic)


def guess_law(main_type: str) -> str:
    return LAW_MAP.get(main_type, "\ub3c5\uc810\uaddc\uc81c \ubc0f \uacf5\uc815\uac70\ub798\uc5d0 \uad00\ud55c \ubc95\ub960")


def choose_persona(main_type: str, sub_type: str, title: str) -> str:
    general_keywords = [
        "\uc804\uc790\uc0c1\uac70\ub798",
        "\ud45c\uc2dc\uad11\uace0",
        "\uace0\uac1d\uc720\uc778",
        "\ubc29\ubb38\ud310\ub9e4",
        "\uac00\ub9f9",
        "\uc18c\ube44\uc790",
    ]
    student_keywords = [
        "\ubd80\ub2f9\ud55c \uacf5\ub3d9\ud589\uc704",
        "\ub2e8\uccb4-\uacbd\uc7c1\uc81c\ud55c\ud589\uc704",
        "\uc0ac\uc5c5\uc790\ub2e8\uccb4\uae08\uc9c0\ud589\uc704",
        "\uacbd\uc81c\ub825 \uc9d1\uc911\uc5b5\uc81c",
    ]
    corpus = " ".join([main_type, sub_type, title])
    if any(keyword in corpus for keyword in general_keywords):
        return PERSONA_GENERAL
    if any(keyword in corpus for keyword in student_keywords) or "\uc785\ucc30" in title or re.search(r"\d+\uac1c \uc0ac\uc5c5\uc790", title):
        return PERSONA_STUDENT
    return PERSONA_COMPLIANCE


def choose_question_type(persona: str, main_type: str, sub_type: str, title: str, actions: list[str]) -> str:
    corpus = " ".join([main_type, sub_type, title, " ".join(actions)])
    if persona == PERSONA_GENERAL:
        if any(keyword in corpus for keyword in ["\ud45c\uc2dc\uad11\uace0", "\uace0\uac1d\uc720\uc778", "\uc18c\ube44\uc790"]):
            return "\uc2dc\uc7a5\uc601\ud5a5"
        if any(keyword in corpus for keyword in ["\uacfc\uc9d5\uae08", "\uace0\ubc1c", "\uc2dc\uc815\uba85\ub839"]):
            return "\uc81c\uc7ac\uc911\uc2ec"
        return "\ubc95\ub839\uc911\uc2ec"
    if persona == PERSONA_COMPLIANCE:
        if any(keyword in corpus for keyword in ["\uacfc\uc9d5\uae08", "\uace0\ubc1c", "\ub300\uae08 \ubbf8\uc9c0\uae09", "\uac10\uc561"]):
            return "\uc81c\uc7ac\uc911\uc2ec"
        if any(keyword in corpus for keyword in ["\uac70\ub798\uc0c1\uc9c0\uc704\ub0a8\uc6a9", "\ubd88\uc774\uc775\uc81c\uacf5", "\uc9c0\uc6d0", "\uad6c\uc18d\uc870\uac74", "\ub300\uae08", "\ud45c\uc2dc\uc0ac\ud56d"]):
            return "\ubc95\ub839\uc911\uc2ec"
        return "\uc720\uc0ac\uc0ac\ub840"
    if any(keyword in corpus for keyword in ["\uc785\ucc30", "\uacf5\ub3d9", "\uac00\uaca9", "\uc810\uc720\uc728", "\uc9d1\uc911"]):
        return "\ud1b5\uacc4\uae30\ubc18"
    if any(keyword in corpus for keyword in ["\uc2dc\uc7a5", "\uacbd\uc7c1", "\uc601\ud5a5"]):
        return "\uc2dc\uc7a5\uc601\ud5a5"
    return "\uc720\uc0ac\uc0ac\ub840"


def choose_difficulty(main_type: str, sub_type: str, title: str, info_count: int, actions: list[str]) -> str:
    corpus = " ".join([main_type, sub_type, title, " ".join(actions)])
    advanced_signals = [
        "\uc785\ucc30\ub2f4\ud569",
        "\uacbd\uc81c\ub825 \uc9d1\uc911\uc5b5\uc81c",
        "\uace0\ubc1c",
        "\ubc95\uc704\ubc18\uacf5\ud45c",
        "\uc9c0\uc8fc\ud68c\uc0ac\uc124\ub9bd,\uc804\ud658",
    ]
    moderate_signals = [
        "\ud558\ub3c4\uae09",
        "\uac00\ub9f9",
        "\uac70\ub798\uc0c1\uc9c0\uc704\ub0a8\uc6a9",
        "\ud45c\uc2dc\uad11\uace0",
        "\uc804\uc790\uc0c1\uac70\ub798",
        "\ub300\uaddc\ubaa8 \uc720\ud1b5",
        "\ubc29\ubb38\ud310\ub9e4",
        "\uace0\uac1d\uc720\uc778",
        "\ub2e8\uccb4-",
        "\uacf5\ub3d9-",
    ]
    if (
        info_count >= 3
        or re.search(r"\d+\uac1c \uc0ac\uc5c5\uc790", title)
        or any(keyword in corpus for keyword in advanced_signals)
    ):
        return "\uc2ec\ud654"
    if any(keyword in corpus for keyword in moderate_signals) or info_count == 2:
        return "\uc911\uae09"
    return "\uae30\ucd08"


def find_chunk(chunks: list[dict], predicates: list[str], chunk_type: str = "text") -> dict | None:
    for chunk in chunks:
        meta = chunk.get("metadata", {})
        joined = " ".join(
            str(meta.get(key, ""))
            for key in ("Header", "Header2", "Header3", "section", "chunk_type")
        )
        if chunk_type and meta.get("chunk_type") != chunk_type:
            continue
        if any(token in joined for token in predicates):
            return chunk
    return None


def first_text_chunk(chunks: list[dict], section_name: str | None = None) -> dict | None:
    for chunk in chunks:
        meta = chunk.get("metadata", {})
        if meta.get("chunk_type") != "text":
            continue
        if section_name and meta.get("section") != section_name:
            continue
        return chunk
    return None


def compose_question(persona: str, question_type: str, subject: str, title: str, actions: list[str]) -> str:
    action_text = ", ".join(actions) if actions else "\uc870\uce58"
    variants = {
        PERSONA_GENERAL: {
            "\uc81c\uc7ac\uc911\uc2ec": [
                "\uc774 \uc0ac\uac74\uc5d0\uc11c \uc5b4\ub5a4 \uc81c\uc7ac\uac00 \ub0b4\ub824\uc84c\ub098\uc694?",
                f"'{title}' \uc0ac\uac74\uc5d0\uc11c \ud655\uc778\ub418\ub294 \uc870\uce58\ub294 \ubb34\uc5c7\uc778\uac00\uc694?",
                f"'{subject}' \ubb38\uc81c\ub85c \uc5b4\ub5a4 \ucc98\ubd84\uc774 \ub0b4\ub824\uc84c\ub098\uc694?",
            ],
            "\ubc95\ub839\uc911\uc2ec": [
                f"'{subject}' \ud589\uc704\ub294 \uc65c \ubb38\uc81c\ub85c \ud310\ub2e8\ub418\ub098\uc694?",
                f"'{title}' \uc0ac\uac74\uc5d0\uc11c \uc704\ubc95\uc131\uc740 \uc5b4\ub5bb\uac8c \uc124\uba85\ub418\ub098\uc694?",
                f"'{subject}' \ud589\uc704\ub97c \uc758\uacb0\uc11c\uac00 \ubb38\uc81c 삼은 \uc774\uc720\ub294 \ubb34\uc5c7\uc778\uac00\uc694?",
            ],
            "\uc720\uc0ac\uc0ac\ub840": [
                f"'{subject}' \uc0ac\uac74\uacfc \ube44\uc2b7\ud55c \uc0ac\ub840\ub97c \ubcfc \ub54c \ud3ec\uc778\ud2b8\ub294 \ubb34\uc5c7\uc778\uac00\uc694?",
                f"'{title}' \uc0ac\uac74\uacfc \uc720\uc0ac \uc0ac\ub840\ub97c \ube44\uad50\ud560 \ub54c \uc5b4\ub514\ub97c \ubd10\uc57c \ud558\ub098\uc694?",
            ],
            "\ud1b5\uacc4\uae30\ubc18": [
                f"'{subject}' \uc0ac\uac74\uc740 \uc5b4\ub5a4 \uc704\ubc18 \uc720\ud615\uc73c\ub85c \ubd84\ub958\ud574 \ubcfc \uc218 \uc788\ub098\uc694?",
                f"'{title}' \uc0ac\uac74\uc740 \uc5b4\ub5a4 \uc720\ud615\uc758 \uc704\ubc18\uc73c\ub85c \uc774\ud574\ud558\uba74 \ub418\ub098\uc694?",
            ],
            "\uc2dc\uc7a5\uc601\ud5a5": [
                f"'{subject}' \ud589\uc704\uac00 \uc18c\ube44\uc790\ub098 \uc2dc\uc7a5\uc5d0 \ubbf8\uce60 \uc601\ud5a5\uc740 \ubb34\uc5c7\uc778\uac00\uc694?",
                f"'{title}' \uc0ac\uac74\uc5d0\uc11c \uc18c\ube44\uc790 \ud53c\ud574 \ub610\ub294 \uc2dc\uc7a5 \uc601\ud5a5\uc740 \uc5b4\ub5bb\uac8c \ubcfc \uc218 \uc788\ub098\uc694?",
            ],
        },
        PERSONA_COMPLIANCE: {
            "\uc81c\uc7ac\uc911\uc2ec": [
                "\uc81c\uc7ac \uc218\uc900\uacfc \uc0b0\uc815\uc5d0 \uc601\ud5a5\uc744 \uc900 \uc2e4\ubb34 \ud3ec\uc778\ud2b8\ub294 \ubb34\uc5c7\uc778\uac00\uc694?",
                f"'{title}' \uc0ac\uac74\uc5d0\uc11c {action_text} \uc218\uc900\uc744 \uc608\uce21\ud558\uac8c \ud558\ub294 \ud3ec\uc778\ud2b8\ub294 \ubb34\uc5c7\uc778\uac00\uc694?",
                f"'{subject}' \uad00\ub828 \uc2e4\ubb34\uc5d0\uc11c \uc81c\uc7ac \ub9ac\uc2a4\ud06c\ub97c \ud0a4\uc6b0\ub294 \uc694\uc18c\ub294 \ubb34\uc5c7\uc778\uac00\uc694?",
            ],
            "\ubc95\ub839\uc911\uc2ec": [
                f"'{subject}' \uad00\ub828 \uc2e4\ubb34\uc5d0\uc11c \uc5b4\ub5a4 \ubc95\ub839 \ub9ac\uc2a4\ud06c\uc640 \ud1b5\uc81c \ud3ec\uc778\ud2b8\ub97c \ubd10\uc57c \ud558\ub098\uc694?",
                f"'{title}' \uc0ac\uc548\uc744 \uae30\uc900\uc73c\ub85c \ubcf4\uba74 \ub0b4\ubd80 \ud1b5\uc81c\uc5d0\uc11c \ubb34\uc5c7\uc744 \uc810\uac80\ud574\uc57c \ud558\ub098\uc694?",
                f"'{subject}' \uc720\ud615\uc740 \uc900\ubc95 \uad00\uc810\uc5d0\uc11c \uc5b4\ub5a4 \uc808\ucc28 \ubbf8\ube44\ub85c \uc774\uc5b4\uc9c0\ub098\uc694?",
            ],
            "\uc720\uc0ac\uc0ac\ub840": [
                f"'{subject}'\uacfc \uc720\uc0ac \uc0ac\uc548\uc744 \ube44\uad50\ud560 \ub54c \uc5b4\ub5a4 \ud1b5\uc81c \uc2e4\ud328\uac00 \ub4dc\ub7ec\ub098\ub098\uc694?",
                f"'{title}' \uc0ac\uac74\uacfc \uc720\uc0ac \uc0ac\ub840\ub97c \ube44\uad50\ud558\uba74 \ubc18\ubcf5\ub418\ub294 \ud1b5\uc81c \ucde8\uc57d\uc810\uc740 \ubb34\uc5c7\uc778\uac00\uc694?",
            ],
            "\ud1b5\uacc4\uae30\ubc18": [
                f"'{subject}' \uc0ac\uac74\uc740 \ub3d9\uc885 \uc704\ubc18\uad70 \uad00\uc810\uc5d0\uc11c \uc5b4\ub5a4 \uc704\ud5d8\uc2e0\ud638\ub85c \ubd84\ub958\ub420 \uc218 \uc788\ub098\uc694?",
                f"'{title}' \uc0ac\uac74\uc744 \ub3d9\uc885 \ub9ac\uc2a4\ud06c \uad70\uc9d1\uc5d0 \ub123\ub294\ub2e4\uba74 \uc5b4\ub5a4 \ubcc0\uc218\ub97c \ubcf4\uba74 \ub418\ub098\uc694?",
            ],
            "\uc2dc\uc7a5\uc601\ud5a5": [
                f"'{subject}' \ud589\uc704\uac00 \uac70\ub798\uc0c1\ub300\ubc29\uacfc \uc2dc\uc7a5\uc5d0 \ubbf8\uce5c \uc601\ud5a5\uc740 \ubb34\uc5c7\uc778\uac00\uc694?",
                f"'{title}' \uc0ac\uac74\uc744 \ud1b5\ud574 \ubcf8 \uac70\ub798\uc0c1\ub300\ubc29 \ubc0f \uc2dc\uc7a5 \uc601\ud5a5\uc740 \ubb34\uc5c7\uc778\uac00\uc694?",
            ],
        },
        PERSONA_STUDENT: {
            "\uc81c\uc7ac\uc911\uc2ec": [
                "\ud574\ub2f9 \uc81c\uc7ac\uc758 \ube44\ub840\uc131\uacfc \ud310\ub2e8 \uad6c\uc870\ub97c \uc5b4\ub5bb\uac8c \ud574\uc11d\ud560 \uc218 \uc788\ub098\uc694?",
                f"'{title}' \uc0ac\uac74\uc5d0\uc11c \uc81c\uc7ac \uc218\uc900\uc740 \uc5b4\ub5a4 \ud310\ub2e8 \uad6c\uc870 \uc18d\uc5d0\uc11c \uc124\uba85\ub420 \uc218 \uc788\ub098\uc694?",
            ],
            "\ubc95\ub839\uc911\uc2ec": [
                f"'{subject}' \ud310\ub2e8\uc758 \ubc95\uc801 \uadfc\uac70\ub97c \uc5b4\ub5bb\uac8c \uc815\ub9ac\ud560 \uc218 \uc788\ub098\uc694?",
                f"'{title}' \uc0ac\uac74\uc5d0\uc11c \ubc95\ub9ac \ud574\uc11d\uc758 \ud575\uc2ec\uc740 \ubb34\uc5c7\uc778\uac00\uc694?",
                f"'{subject}' \ud589\uc704\uc5d0 \ub300\ud55c \ud310\ub2e8 \uadfc\uac70\ub97c \uc5f0\uad6c\uc6a9\uc73c\ub85c \uc815\ub9ac\ud558\ub824\uba74 \uc5b4\ub5bb\uac8c \ubcf4\uba74 \ub418\ub098\uc694?",
            ],
            "\uc720\uc0ac\uc0ac\ub840": [
                f"'{subject}'\uacfc \uc720\uc0ac \uc0ac\ub840\ub97c \ube44\uad50\ud560 \ub54c \ucc28\ubcc4\uc810\uc740 \ubb34\uc5c7\uc778\uac00\uc694?",
                f"'{title}' \uc0ac\uac74\uacfc \uc720\uc0ac \ud310\ub2e8\ub840\ub97c \ube44\uad50\ud558\ub824\uba74 \uc5b4\ub5a4 \ucd95\uc744 \uc138\uc6cc\uc57c \ud558\ub098\uc694?",
            ],
            "\ud1b5\uacc4\uae30\ubc18": [
                f"'{subject}' \uc0ac\uac74\uc744 \ub3d9\uc885 \uc0ac\uac74\uad70\uacfc \ube44\uad50\ud558\ub824\uba74 \uc5b4\ub5a4 \ubd84\uc11d \ubcc0\uc218\uac00 \ud544\uc694\ud55c\uac00\uc694?",
                f"'{title}' \uc0ac\uac74\uc744 \ud1b5\uacc4\uc801\uc73c\ub85c \ube44\uad50\ud558\ub824\uba74 \uc5b4\ub5a4 \ubd84\uc11d \ub2e8\uc704\uac00 \ud544\uc694\ud55c\uac00\uc694?",
                f"'{subject}' \uc720\ud615\uc744 \uc5f0\uad6c \ubcc0\uc218\ub85c \uc0bc\uc744 \ub54c \uc5b4\ub5a4 \ube44\uad50 \uc9c0\ud45c\uac00 \uc720\uc6a9\ud55c\uac00\uc694?",
            ],
            "\uc2dc\uc7a5\uc601\ud5a5": [
                f"'{subject}' \ud589\uc704\uc758 \uc2dc\uc7a5\ud6a8\uacfc\uc640 \uacbd\uc7c1\uc81c\ud55c\uc131\uc740 \uc5b4\ub5bb\uac8c \uc124\uba85\ud560 \uc218 \uc788\ub098\uc694?",
                f"'{title}' \uc0ac\uac74\uc5d0\uc11c \uc2dc\uc7a5 \uc601\ud5a5\uc744 \uc5f0\uad6c\uc801\uc73c\ub85c \uc815\ub9ac\ud558\ub824\uba74 \ubb34\uc5c7\uc744 \ubd10\uc57c \ud558\ub098\uc694?",
            ],
        },
    }
    choices = variants[persona][question_type]
    return choices[stable_variant(title, subject, action_text, modulo=len(choices))]


def compose_answer(persona: str, question_type: str, subject: str, parties: list[str], actions: list[str], main_type: str, fact_text: str, law_name: str) -> str:
    actor = parties[0] if parties else "\ud53c\uc2ec\uc778"
    action_text = ", ".join(actions) if actions else "\uc2dc\uc815\uc870\uce58"
    fact_summary = shorten(fact_text, 110) if fact_text else f"'{subject}'\uc774 \ud575\uc2ec \uc7c1\uc810\uc774\ubbc0\ub85c"
    party_summary = f"\ud53c\uc2ec\uc778 {len(parties)}\uac1c \uc8fc\uccb4" if len(parties) > 1 else actor
    if persona == PERSONA_GENERAL:
        if question_type == "\uc2dc\uc7a5\uc601\ud5a5":
            variants = [
                f"\uc774 \uc0ac\uac74\uc740 '{subject}'\uc640 \uad00\ub828\ub41c \ud589\uc704\uac00 \uc18c\ube44\uc790 \uc120\ud0dd\uacfc \uacf5\uc815\ud55c \uac70\ub798 \uc9c8\uc11c\uc5d0 \ubd80\uc815\uc801 \uc601\ud5a5\uc744 \uc904 \uc218 \uc788\ub2e4\ub294 \uc810\uc5d0\uc11c \ubb38\uc81c \ub418\uc5c8\uc2b5\ub2c8\ub2e4. \ud2b9\ud788 {fact_summary} \uc810\uc774 \ud310\ub2e8\uc758 \ubc30\uacbd\uc774 \ub418\uc5c8\uc2b5\ub2c8\ub2e4.",
                f"\uc758\uacb0\uc11c\ub294 {party_summary}\uc758 '{subject}' \ud589\uc704\uac00 \uc2dc\uc7a5 \uc2e0\ub8b0\uc640 \uc18c\ube44\uc790 \uc120\ud0dd\uc744 \ud574\uce60 \uc218 \uc788\ub2e4\uace0 \ubd24\uc2b5\ub2c8\ub2e4. \ud575\uc2ec \ubc30\uacbd\uc740 {fact_summary} \ub0b4\uc6a9\uc785\ub2c8\ub2e4.",
            ]
            return variants[stable_variant(subject, action_text, modulo=len(variants))]
        if question_type == "\ubc95\ub839\uc911\uc2ec":
            variants = [
                f"\uc758\uacb0\uc11c\ub294 {actor}\uc758 \ud589\uc704\ub97c '{subject}' \uad00\ub828 \uc704\ubc18\uc73c\ub85c \ubcf4\uace0, {law_name} \uad00\ub828 \uaddc\uc815 \uc801\uc6a9 \uac00\ub2a5\uc131\uc744 \uc804\uc81c\ub85c \ud310\ub2e8\ud588\uc2b5\ub2c8\ub2e4. \ud575\uc2ec\uc740 \uac70\ub798 \uc0c1\ub300\ubc29 \ub610\ub294 \uc18c\ube44\uc790\uc5d0\uac8c \ubd88\uc774\uc775\uc744 \uc8fc\ub294 \uad6c\uc870\uac00 \uc788\uc5c8\ub294\uc9c0\uc785\ub2c8\ub2e4.",
                f"'{subject}' \ud589\uc704\ub294 {law_name} \uad00\ub828 \uaddc\uc815\uc774 \uac00\ub2a5\ud55c \uc0ac\uc548\uc73c\ub85c \uc815\ub9ac\ub429\ub2c8\ub2e4. \uc758\uacb0\uc11c\ub294 {fact_summary} \uc0ac\uc815\uc744 \ubc14\ud0d5\uc73c\ub85c \uc704\ubc95\uc131\uc744 \ud310\ub2e8\ud588\uc2b5\ub2c8\ub2e4.",
            ]
            return variants[stable_variant(subject, actor, modulo=len(variants))]
        if question_type == "\ud1b5\uacc4\uae30\ubc18":
            return f"\uc774 \uc0ac\uac74\uc740 '{main_type}' \ubc94\uc8fc \uc548\uc5d0\uc11c '{subject}'\uc73c\ub85c \ubd84\ub958\ud560 \uc218 \uc788\uc2b5\ub2c8\ub2e4. \ube44\uad50 \uc2dc\uc5d0\ub294 \uc870\uce58 \uc720\ud615({action_text}), \ud53c\uc2ec\uc778 \uc218, \ud53c\ud574 \ubc94\uc704, \ud589\uc704 \ubc29\uc2dd\uc744 \ud568\uaed8 \ubcf4\uba74 \uc720\uc0ac \uc0ac\uac74 \uad6c\ubd84\uc5d0 \ub3c4\uc6c0\uc774 \ub429\ub2c8\ub2e4."
        if question_type == "\uc720\uc0ac\uc0ac\ub840":
            return f"\ube44\uc2b7\ud55c \uc0ac\uac74\uacfc \ube44\uad50\ud560 \ub54c\ub294 '{subject}'\uc774 \uc5b4\ub5a4 \uac70\ub798 \uad6c\uc870\uc5d0\uc11c \ubc1c\uc0dd\ud588\ub294\uc9c0, \uadf8 \uacb0\uacfc {action_text} \uc870\uce58\uac00 \ubcd1\uacfc\ub418\uc5c8\ub294\uc9c0\ub97c \ubcf4\uba74 \ub429\ub2c8\ub2e4. \ud2b9\ud788 {fact_summary} \uc0ac\uc815\uc774 \ube44\uad50 \ud3ec\uc778\ud2b8\uc785\ub2c8\ub2e4."
        return f"\uc774 \uc0ac\uac74\uc5d0\uc11c \ud574\ub2f9 \uc0ac\uc5c5\uc790\ub294 '{subject}' \uad00\ub828 \ud589\uc704\ub85c {action_text} \uc870\uce58\ub97c \ubc1b\uc558\uc2b5\ub2c8\ub2e4. \uc758\uacb0\uc11c\ub294 \ud589\uc704\uac00 {main_type} \ubc94\uc8fc\uc5d0 \ud574\ub2f9\ud55c\ub2e4\uace0 \ubcf4\uace0 \uac70\ub798 \uc9c8\uc11c \uc800\ud574 \uac00\ub2a5\uc131\uc744 \ubb38\uc81c \uc0bc\uc558\uc2b5\ub2c8\ub2e4."

    if persona == PERSONA_COMPLIANCE:
        if question_type == "\ubc95\ub839\uc911\uc2ec":
            variants = [
                f"\uc2e4\ubb34\uc0c1\uc73c\ub85c\ub294 '{subject}'\uac00 {law_name} \uc0c1 \uae08\uc9c0 \ub610\ub294 \uc758\ubb34 \uc704\ubc18\uc5d0 \uc5f0\uacb0\ub420 \uc218 \uc788\ub294\uc9c0\ub97c \uc6b0\uc120 \ubd10\uc57c \ud569\ub2c8\ub2e4. \ud2b9\ud788 {fact_summary} \uc0ac\uc2e4\uc774 \uc0ac\uc804 \uc2b9\uc778, \uc11c\uba74\ud654, \uac70\ub798\uc0c1\ub300\ubc29 \ub3d9\uc758 \uc808\ucc28\uac00 \uc81c\ub300\ub85c \uc791\ub3d9\ud588\ub294\uc9c0 \uc810\uac80\ud560 \ud3ec\uc778\ud2b8\uc785\ub2c8\ub2e4.",
                f"'{subject}' \uc0ac\uc548\uc740 {law_name} \uc0c1 \uae08\uc9c0\ud589\uc704 \ud574\ub2f9\uc131\ucac4 \uc808\ucc28 \ud1b5\uc81c \ubbf8\ube44 \uc5ec\ubd80\ub97c \ud568\uaed8 \ubd10\uc57c \ud569\ub2c8\ub2e4. \uc774 \uc0ac\uac74\uc5d0\uc11c\ub294 {fact_summary} \uc0ac\uc815\uc774 \ud575\uc2ec \ud1b5\uc81c \ud3ec\uc778\ud2b8\ub85c \ub4dc\ub7ec\ub0a9\ub2c8\ub2e4.",
            ]
            return variants[stable_variant(subject, actor, modulo=len(variants))]
        if question_type == "\uc720\uc0ac\uc0ac\ub840":
            return f"\uc720\uc0ac \uc0ac\uc548\uacfc \ube44\uad50\ud558\uba74 \uc774 \uc0ac\uac74\uc740 '{subject}'\uac00 \ub0b4\ubd80 \ud1b5\uc81c \ubd80\uc7ac \ub610\ub294 \uc6b0\uc6d4\uc801 \uc9c0\uc704 \ud589\uc0ac \ubb38\uc81c\ub85c \uc5f0\uacb0\ub418\uc5c8\ub2e4\ub294 \uc810\uc774 \ub4dc\ub7ec\ub0a9\ub2c8\ub2e4. \ud2b9\ud788 {fact_summary} \ub0b4\uc6a9\uc774 \uc7ac\ubc1c\ubc29\uc9c0 \uccb4\uacc4\ub97c \uc124\uacc4\ud560 \ub54c \ube44\uad50 \uae30\uc900\uc774 \ub429\ub2c8\ub2e4."
        if question_type == "\ud1b5\uacc4\uae30\ubc18":
            return f"\ub3d9\uc885 \uc704\ubc18\uad70 \uad00\uc810\uc5d0\uc11c\ub294 '{subject}'\uc744 \ub9ac\uc2a4\ud06c \uc2e0\ud638\ub85c \ubd84\ub958\ud560 \uc218 \uc788\uc2b5\ub2c8\ub2e4. \uc2e4\ubb34\uc5d0\uc11c\ub294 \uc704\ubc18 \uc720\ud615, \uc870\uce58 \uc218\uc900, \ud53c\ud574\uc790 \ubc94\uc704, \uc7ac\ubc1c\uc131\uc744 \ubcc0\uc218\ub85c \ub450\uace0 \ube44\uad50\ud558\uba74 \ud1b5\uc81c \uc6b0\uc120\uc21c\uc704\ub97c \uc815\ud558\uae30 \uc218\uc6d4\ud569\ub2c8\ub2e4."
        if question_type == "\uc2dc\uc7a5\uc601\ud5a5":
            return f"\uc774 \ud589\uc704\ub294 \uac70\ub798\uc0c1\ub300\ubc29\uc758 \ud611\uc0c1\ub825\uc744 \uc57d\ud654\uc2dc\ud0a4\uace0 \uc2dc\uc7a5 \uc2e0\ub8b0\ub97c \uc800\ud574\ud560 \uc218 \uc788\uc2b5\ub2c8\ub2e4. \ud1b5\uc81c \uad00\uc810\uc5d0\uc11c\ub294 \ub300\uc678 \uace0\uc9c0, \ube44\uc6a9 \ubd84\ub2f4, \uc870\uac74 \ubcc0\uacbd \uc2b9\uc778 \uccb4\uacc4\ub97c \ud568\uaed8 \uc810\uac80\ud558\ub294 \uac83\uc774 \ud544\uc694\ud569\ub2c8\ub2e4."
        variants = [
            f"\uc81c\uc7ac \uc218\uc900\uc740 {action_text} \uc870\uce58\uc5d0\uc11c \ud655\uc778\ub429\ub2c8\ub2e4. \uc2e4\ubb34\uc5d0\uc11c\ub294 '{subject}' \ud589\uc704\uac00 \ubc1c\uc0dd\ud55c \uacbd\uc704, \ubd80\ub2f9\uc131 \ud310\ub2e8 \uadfc\uac70, \uc7ac\ubc1c\uc131 \uc5ec\ubd80\ub97c \ud568\uaed8 \ubd10\uc57c \uc81c\uc7ac \uc218\uc900\uc744 \uc608\uce21\ud560 \uc218 \uc788\uc2b5\ub2c8\ub2e4.",
            f"\uc774 \uc0ac\uac74\uc5d0\uc11c \ud655\uc778\ub418\ub294 \uc870\uce58\ub294 {action_text}\uc785\ub2c8\ub2e4. \uc81c\uc7ac \uc2e4\ubb34\ub97c \ubcf4\ub824\uba74 {fact_summary} \uc0ac\uc815\uacfc \uac19\uc774 \uc704\ubc95\uc131 \uadfc\uac70\uac00 \uc5b4\ub5bb\uac8c \ud615\uc131\ub410\ub294\uc9c0 \ud568\uaed8 \ubd10\uc57c \ud569\ub2c8\ub2e4.",
            f"{action_text} \uc870\uce58\uac00 \ub0b4\ub824\uc84c\ub2e4\ub294 \uc810\uc740 '{subject}' \ud589\uc704\uc758 \uc704\ubc95\uc131\uc744 \uc2e4\ubb34\uc0c1 \ubb34\uac81\uac8c \ubd24\ub2e4\ub294 \ub73b\uc785\ub2c8\ub2e4. \uc774\ub54c\ub294 \ubc1c\uc0dd \uacbd\uc704, \uac70\ub798\uc0c1\ub300\ubc29 \ud53c\ud574, \uc7ac\ubc1c \uac00\ub2a5\uc131\uc744 \ud568\uaed8 \uc810\uac80\ud574\uc57c \ud569\ub2c8\ub2e4.",
        ]
        return variants[stable_variant(subject, action_text, actor, modulo=len(variants))]

    if question_type == "\ubc95\ub839\uc911\uc2ec":
        return f"\uc774 \uc0ac\uac74\uc740 '{subject}'\uc774 {law_name} \ud574\uc11d \ubb38\uc81c\uc640 \uc5b4\ub5bb\uac8c \uc5f0\uacb0\ub418\ub294\uc9c0 \ubcf4\uc5ec\uc90d\ub2c8\ub2e4. \ud2b9\ud788 {fact_summary} \uc0ac\uc2e4\uad00\uacc4\ub97c \uae30\ubc18\uc73c\ub85c \ud589\uc704\uc720\ud615, \uacbd\uc7c1\uc81c\ud55c\uc131, \ube44\ub840\uc131 \ud310\ub2e8\uc744 \uad6c\uc870\ud654\ud560 \uc218 \uc788\uc2b5\ub2c8\ub2e4."
    if question_type == "\ud1b5\uacc4\uae30\ubc18":
        variants = [
            f"\ube44\uad50\uc5f0\uad6c \uad00\uc810\uc5d0\uc11c\ub294 '{subject}'\uc744 \uc8fc\uc694 \ubd84\uc11d \ub2e8\uc704\ub85c \uc0bc\uc744 \uc218 \uc788\uc2b5\ub2c8\ub2e4. \uc0ac\uac74 \uc218, \ud53c\uc2ec\uc778 \uc218({len(parties)}), \uc870\uce58 \uc720\ud615({action_text}), \uacbd\uc7c1\uc81c\ud55c \ud6a8\uacfc, \ud53c\ud574 \uc2dc\uc7a5 \ubc94\uc704\ub97c \ubcc0\uc218\ud654\ud558\uba74 \ub3d9\uc885 \uc0ac\uac74\uad70\uacfc \ube44\uad50\uac00 \uac00\ub2a5\ud569\ub2c8\ub2e4.",
            f"'{subject}' \uc720\ud615\uc740 \ube44\uad50 \ubd84\uc11d \uc2dc \ud53c\uc2ec\uc778 \uad6c\uc131, \uc870\uce58 \uc218\uc900({action_text}), \uc2dc\uc7a5 \ubc94\uc704, \uc99d\uac70 \uad6c\uc870\ub97c \ud568\uaed8 \ubcf4\ub294 \uac83\uc774 \uc720\uc6a9\ud569\ub2c8\ub2e4. \ud2b9\ud788 {fact_summary} \uc0ac\uc815\uc740 \uc0ac\uac74\uad70 \ube44\uad50\uc5d0 \ud65c\uc6a9\ud560 \uc218 \uc788\uc2b5\ub2c8\ub2e4.",
        ]
        return variants[stable_variant(subject, action_text, fact_summary, modulo=len(variants))]
    if question_type == "\uc2dc\uc7a5\uc601\ud5a5":
        return f"\ud559\uc220\uc801\uc73c\ub85c\ub294 '{subject}'\uc774 \uac70\ub798 \uc0c1\ub300\ubc29 \ub610\ub294 \uc18c\ube44\uc790 \ud6c4\uc0dd\uc5d0 \uc5b4\ub5a4 \ube44\uc6a9\uc744 \uc720\ubc1c\ud588\ub294\uc9c0 \uc124\uba85\ud560 \uc218 \uc788\uc2b5\ub2c8\ub2e4. \uc774 \uc0ac\uac74\uc740 \ud589\uc704 \uad6c\uc870\ub97c \ud1b5\ud574 \uc2dc\uc7a5 \uc9c8\uc11c \uc800\ud574 \uc591\uc0c1\uc744 \uad00\ucc30\ud560 \uc218 \uc788\ub294 \uc0ac\ub840\uc785\ub2c8\ub2e4."
    if question_type == "\uc720\uc0ac\uc0ac\ub840":
        return f"\uc720\uc0ac \uc0ac\ub840\uc640 \ube44\uad50\ud558\uba74 \uc774 \uc0ac\uac74\uc740 '{subject}'\uac00 \ub4dc\ub7ec\ub098\ub294 \ubc29\uc2dd, \uc99d\uac70 \uad6c\uc870, \uc870\uce58 \uc218\uc900\uc5d0\uc11c \ud2b9\uc9d5\uc744 \uac00\uc9d1\ub2c8\ub2e4. \ube44\uad50 \ucd95\uc73c\ub85c\ub294 \uc704\ubc18 \uc720\ud615, \uc2dc\uc7a5 \ubc94\uc704, \uc870\uce58 \uacb0\uacfc, \ud310\ub2e8 \ubb38\uad6c\ub97c \uc0bc\uc744 \uc218 \uc788\uc2b5\ub2c8\ub2e4."
    return f"\uc81c\uc7ac\ub294 {action_text}\uc73c\ub85c \uc815\ub9ac\ub418\uba70, \uc774\ub294 '{subject}'\uc5d0 \ub300\ud55c \uc704\ubc95\uc131 \ud310\ub2e8\uacfc \ube44\ub840\uc131 \uac80\ud1a0\uc758 \uacb0\uacfc\ub85c \ud574\uc11d\ud560 \uc218 \uc788\uc2b5\ub2c8\ub2e4."


def compose_summary(title: str, subject: str, main_type: str, sub_type: str, parties: list[str], actions: list[str]) -> str:
    actor = parties[0] if parties else "\ud53c\uc2ec\uc778"
    action_text = ", ".join(actions) if actions else "\uc2dc\uc815\uc870\uce58"
    detail = clean_subject(sub_type or main_type or subject)
    return f"{title} \uc0ac\uac74\uc740 {actor}\uc758 '{detail}' \ud589\uc704\uac00 {main_type} \ubc94\uc8fc\uc5d0\uc11c \ubb38\uc81c \ub41c \uc0ac\uc548\uc73c\ub85c, \uc758\uacb0\uc11c\ub294 {action_text} \uc870\uce58\ub97c \ud1b5\ud574 \uac70\ub798 \uc9c8\uc11c \ud68c\ubcf5 \ud544\uc694\uc131\uc744 \uc81c\uc2dc\ud55c\ub2e4."


def compose_facts(parties: list[str], main_type: str, subject: str, fact_text: str) -> str:
    actor = parties[0] if parties else "\ud53c\uc2ec\uc778"
    if fact_text:
        return f"\uc8fc\uccb4: {actor} / \ud575\uc2ec \uc720\ud615: {main_type} / \uc8fc\uc694 \uc0ac\uc2e4: {shorten(fact_text, 180)}"
    return f"\uc8fc\uccb4: {actor} / \ud575\uc2ec \uc720\ud615: {main_type} / \ud575\uc2ec \uc7c1\uc810: {subject}"


def compose_legal(main_type: str, sub_type: str, law_name: str, law_text: str) -> str:
    lead = clean_subject(sub_type or main_type)
    if law_text:
        return f"\ubc95\ub9ac \ud574\uc11d: '{lead}'\uc740(\ub294) {law_name} \uc0c1 \ud589\uc704\uc720\ud615 \ud574\ub2f9\uc131, \uac70\ub798\uc0c1 \uc6b0\uc704 \ub610\ub294 \uacbd\uc7c1\uc81c\ud55c\uc131, \uc0c1\ub300\ubc29 \ubd88\uc774\uc775 \ubc1c\uc0dd \uc5ec\ubd80\ub97c \uc911\uc2ec\uc73c\ub85c \ud310\ub2e8\ub420 \uc218 \uc788\uc73c\uba70, \uc758\uacb0\uc11c\ub294 {shorten(law_text, 180)}\ub77c\ub294 \uadfc\uac70\ub97c \uc81c\uc2dc\ud55c\ub2e4."
    return f"\ubc95\ub9ac \ud574\uc11d: '{lead}'\uc740(\ub294) {law_name} \uc0c1 \uae08\uc9c0\ud589\uc704 \ub610\ub294 \uc758\ubb34 \uc704\ubc18 \ud310\ub2e8 \uae30\uc900\uacfc \uc5f0\ub3d9\ub418\ub294 \ucabd\uc73c\ub85c \ud574\uc11d\ub420 \uc218 \uc788\ub2e4."


def compose_scenarios(persona: str, subject: str, question_type: str) -> tuple[str, str, str, str]:
    if persona == PERSONA_GENERAL:
        return (
            f"\uc0ac\uc6a9\uc790\uac00 '{subject}' \uad00\ub828 \uad81\uae08\uc99d\uc744 \uc790\uc5f0\uc5b4\ub85c \uc785\ub825\ud558\uba74, AI\uac00 \uc26c\uc6b4 \uc124\uba85\uacfc \ud575\uc2ec \uadfc\uac70\ub97c \ud568\uaed8 \uc81c\uc2dc\ud55c\ub2e4.",
            "\ucd94\uac00 \uc2dc\ub098\ub9ac\uc624: \uc720\uc0ac \uc0ac\uac74\uc744 \ud568\uaed8 \uc81c\uc2dc\ud574 \uc18c\ube44\uc790 \ub610\ub294 \uc18c\uc0c1\uacf5\uc778\uc774 \ube44\uad50\ud574 \ubcfc \uc218 \uc788\uac8c \ud55c\ub2e4.",
            "\uc77c\uc0c1 \uc0ac\ub840 \uc5f0\uacb0, \uc5b4\ub824\uc6b4 \ubc95\ub960\uc6a9\uc5b4 \ucd95\uc18c, \uc81c\uc7ac\uc640 \uc601\ud5a5\uc744 \ubc14\ub85c \uc124\uba85",
            "\uc27d\uace0 \uc9c1\uad00\uc801\uc778 \uc124\uba85",
        )
    if persona == PERSONA_COMPLIANCE:
        focus = "\ubc95\ub839 \ud1b5\uc81c \ud3ec\uc778\ud2b8" if question_type == "\ubc95\ub839\uc911\uc2ec" else "\uc81c\uc7ac \uc608\ubc29\uacfc \uc7ac\ubc1c\ubc29\uc9c0 \ud3ec\uc778\ud2b8"
        return (
            f"\uc900\ubc95 \ub2f4\ub2f9\uc790\uac00 '{subject}' \uc0ac\uc548\uc744 \ud1b5\ud574 \ub0b4\ubd80 \ud504\ub85c\uc138\uc2a4 \ub9ac\uc2a4\ud06c\ub97c \uc810\uac80\ud558\ub3c4\ub85d {focus}\ub97c \uc911\uc2ec\uc73c\ub85c \ub2f5\ubcc0\ud55c\ub2e4.",
            "\ucd94\uac00 \uc2dc\ub098\ub9ac\uc624: \uc720\uc0ac \ubc95\ub839 \uc870\ud56d, \uccb4\ud06c\ub9ac\uc2a4\ud2b8, \uc7ac\ubc1c\ubc29\uc9c0 \uc870\uce58\ub97c \ud568\uaed8 \uc815\ub9ac\ud55c\ub2e4.",
            "\uc6b0\uc120 \ud1b5\uc81c \ud3ec\uc778\ud2b8, \uc99d\ube59 \uc808\ucc28, \uc0ac\uc804 \uc2b9\uc778 \ud544\uc694 \uc5ec\ubd80\ub97c \uba85\ud655\ud788 \uc81c\uc2dc",
            "\ub9ac\uc2a4\ud06c \ubc0f \uc900\ubc95 \uc911\uc2ec \uc124\uba85",
        )
    return (
        f"\uc5f0\uad6c\uc790\uac00 '{subject}'\uc744 \ube44\uad50 \ubd84\uc11d \ub610\ub294 \ubb38\ud5cc\uc815\ub9ac \ub300\uc0c1\uc73c\ub85c \uc0bc\uc744 \uc218 \uc788\ub3c4\ub85d \ud310\ub2e8 \uad6c\uc870\ub97c \uc5f0\uad6c\uc6a9 \uc5b8\uc5b4\ub85c \uc815\ub9ac\ud55c\ub2e4.",
        "\ucd94\uac00 \uc2dc\ub098\ub9ac\uc624: \ub3d9\uc885 \uc0ac\uac74\uad70, \uc2dc\uc7a5\ud6a8\uacfc, \uc81c\uc7ac \uc218\uc900\uc744 \ud568\uaed8 \ube44\uad50\ud560 \uc218 \uc788\ub3c4\ub85d \uc5f0\uad6c \ubcc0\uc218\ub97c \uc81c\uc548\ud55c\ub2e4.",
        "\uc8fc\uc694 \uc7c1\uc810, \ud310\ub2e8 \uad6c\uc870, \ube44\uad50 \uac00\ub2a5 \ubcc0\uc218, \uc5f0\uad6c \ud65c\uc6a9\uc131 \uac15\uc870",
        "\ud559\uc220\u00b7\uc5f0\uad6c \uc911\uc2ec \uc124\uba85",
    )


def persona_note(persona: str, refs: str) -> str:
    if persona == PERSONA_GENERAL:
        return f"\uc77c\ubc18 \uc548\ub0b4\uc6a9\uc774\ubbc0\ub85c \uad6c\uccb4 \uc0ac\uc2e4\uad00\uacc4\uc5d0 \ub530\ub77c \uacb0\ub860\uc774 \ub2ec\ub77c\uc9c8 \uc218 \uc788\uc2b5\ub2c8\ub2e4. {refs}"
    if persona == PERSONA_COMPLIANCE:
        return f"\ub0b4\ubd80 \uc758\uc0ac\uacb0\uc815 \uc804 \uc6d0\ubb38\uacfc \uad00\ub828 \uaddc\uc815 \ub300\uc870\uac00 \ud544\uc694\ud569\ub2c8\ub2e4. {refs}"
    return f"\uc5f0\uad6c \uba54\ubaa8: \ube44\uad50 \uc0ac\ub840 \uc218\uc9d1 \uc2dc \ub3d9\uc77c \uc720\ud615\u00b7\uc870\uce58 \uc0ac\uac74\uad70\uacfc \ud568\uaed8 \uac80\ud1a0. {refs}"


def load_documents(base_dir: Path) -> list[dict]:
    data_dir = next(path for path in base_dir.iterdir() if path.is_dir() and path.name.startswith("AI"))
    records: list[dict] = []
    for meta_path in sorted(data_dir.iterdir(), key=lambda path: path.name):
        if not meta_path.name.endswith("_metadata.json"):
            continue
        meta = json.loads(meta_path.read_text(encoding="utf-8-sig"))
        hybrid_path = meta_path.with_name(meta_path.name.replace("_metadata.json", "_hybrid.json"))
        chunks = json.loads(hybrid_path.read_text(encoding="utf-8-sig")) if hybrid_path.exists() else []
        infos = meta.get(K_INFOS) or []
        main_types = uniq([clean_subject(info.get(K_MAIN, "")) for info in infos])
        sub_types = uniq([clean_subject(info.get(K_SUB, "")) for info in infos])
        actions = uniq([clean_subject(info.get(K_ACTION, "")) for info in infos])
        parties = uniq([clean_subject(info.get(K_COMPANY, "")) for info in infos])
        main_type = main_types[0] if main_types else "\uacf5\uc815\uac70\ub798 \uad00\ub828 \uc704\ubc18"
        sub_type = sub_types[0] if sub_types else main_type
        title = clean_subject(meta.get(K_TITLE, meta_path.stem.replace("_metadata", "")))
        persona = choose_persona(main_type, sub_type, title)
        question_type = choose_question_type(persona, main_type, sub_type, title, actions)
        difficulty = choose_difficulty(main_type, sub_type, title, len(infos), actions)

        order_chunk = find_chunk(chunks, ["\uc8fc\ubb38"])
        facts_chunk = find_chunk(chunks, ["\uae30\ucd08\uc0ac\uc2e4", "\uc0ac\uc2e4\uad00\uacc4", "\uc77c\ubc18\ud604\ud669"]) or first_text_chunk(chunks, "\uc774\uc720")
        law_chunk = find_chunk(chunks, ["\uc704\ubc95", "\ud310\ub2e8", "\uad00\ub828 \ubc95", "\ubc95 \uc704\ubc18", "\uac80\ud1a0"]) or first_text_chunk(chunks)

        subject = clean_subject(sub_type)
        if not subject or subject == "\uc704\ubc18\uc720\ud615(\ucd1d\uad04)":
            subject = title_topic(title)

        records.append(
            {
                "doc_id": meta.get(K_DOC_ID, ""),
                "title": title,
                "date": meta.get(K_DATE, ""),
                "persona": persona,
                "question_type": question_type,
                "difficulty": difficulty,
                "subject": subject,
                "main_type": main_type,
                "sub_type": sub_type,
                "actions": actions,
                "parties": parties,
                "law_name": guess_law(main_type),
                "order_chunk": order_chunk,
                "facts_chunk": facts_chunk,
                "law_chunk": law_chunk,
            }
        )
    records.sort(key=lambda item: (item["date"], item["title"]))
    return records


def build_row(index: int, record: dict) -> list[str]:
    refs = []
    if record["doc_id"]:
        refs.append(f"\uc758\uacb0\uc11c\uad00\ub9ac\ubc88\ud638 {record['doc_id']}")
    for label, chunk in [("\uc8fc\ubb38", record["order_chunk"]), ("\uc0ac\uc2e4", record["facts_chunk"]), ("\ud310\ub2e8", record["law_chunk"])]:
        if chunk:
            refs.append(f"{label} {chunk['metadata'].get('chunk_id', '')}")
    ref_text = " / ".join(refs)

    fact_text = normalize_text(record["facts_chunk"]["page_content"]) if record["facts_chunk"] else ""
    law_text = normalize_text(record["law_chunk"]["page_content"]) if record["law_chunk"] else ""
    question = compose_question(
        record["persona"], record["question_type"], record["subject"], record["title"], record["actions"]
    )
    answer = compose_answer(
        record["persona"],
        record["question_type"],
        record["subject"],
        record["parties"],
        record["actions"],
        record["main_type"],
        fact_text,
        record["law_name"],
    )
    summary = compose_summary(
        record["title"],
        record["subject"],
        record["main_type"],
        record["sub_type"],
        record["parties"],
        record["actions"],
    )
    facts = compose_facts(record["parties"], record["main_type"], record["subject"], fact_text)
    legal = compose_legal(record["main_type"], record["sub_type"], record["law_name"], law_text)
    base_scenario, extra_scenario, design_point, style = compose_scenarios(
        record["persona"], record["subject"], record["question_type"]
    )
    action_summary = ", ".join(record["actions"]) if record["actions"] else "\uc2dc\uc815\uc870\uce58"
    evidence = (
        f"\uc758\uacb0\uc11c\ub294 '{record['subject']}' \ud589\uc704\ub97c {record['main_type']} \ubc94\uc8fc\ub85c \ubcf4\uace0 "
        f"{action_summary} \uc870\uce58\ub97c \uc815\ub9ac\ud588\ub2e4."
    )
    note = persona_note(record["persona"], ref_text)

    return [
        f"QA-{index:04d}",
        record["persona"],
        record["question_type"],
        record["difficulty"],
        record["subject"],
        question,
        answer,
        evidence,
        summary,
        facts,
        legal,
        base_scenario,
        extra_scenario,
        design_point,
        style,
        note,
    ]


def reset_sheet_rows(ws, start_row: int = 3) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def fill_sheet(ws, rows: list[list[str]]) -> None:
    reset_sheet_rows(ws, 3)
    for row in rows:
        ws.append(row)


def dump_workbook_json(wb, output_path: Path) -> None:
    payload = {}
    for ws in wb.worksheets:
        payload[ws.title] = [
            [cell for cell in row]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True)
        ]
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    base_dir = Path(__file__).resolve().parent.parent
    workbook_path = next((base_dir / "rag_system").glob("*.xlsx"))
    records = load_documents(base_dir)
    all_rows = [build_row(index, record) for index, record in enumerate(records, start=1)]

    wb = openpyxl.load_workbook(workbook_path)
    ws_all = wb[wb.sheetnames[0]]
    ws_general = wb[wb.sheetnames[1]]
    ws_compliance = wb[wb.sheetnames[2]]
    ws_student = wb[wb.sheetnames[3]]
    ws_dashboard = wb[wb.sheetnames[4]]

    general_rows = [row for row in all_rows if row[1] == PERSONA_GENERAL]
    compliance_rows = [row for row in all_rows if row[1] == PERSONA_COMPLIANCE]
    student_rows = [row for row in all_rows if row[1] == PERSONA_STUDENT]

    ws_all["A1"] = f"\ud398\ub974\uc18c\ub098 \uae30\ubc18 QA\u00b7\uadfc\uac70 \ub370\uc774\ud130\uc14b ({len(all_rows)}\uac74)"
    ws_general["A1"] = f"{PERSONA_GENERAL} \uc804\uc6a9 QA ({len(general_rows)}\uac74)"
    ws_compliance["A1"] = f"{PERSONA_COMPLIANCE} \uc804\uc6a9 QA ({len(compliance_rows)}\uac74)"
    ws_student["A1"] = f"{PERSONA_STUDENT} \uc804\uc6a9 QA ({len(student_rows)}\uac74)"

    fill_sheet(ws_all, all_rows)
    fill_sheet(ws_general, general_rows)
    fill_sheet(ws_compliance, compliance_rows)
    fill_sheet(ws_student, student_rows)

    q_counter = Counter(row[2] for row in all_rows)
    d_counter = Counter(row[3] for row in all_rows)

    ws_dashboard["A1"] = f"\ud398\ub974\uc18c\ub098 \uae30\ubc18 QA \ub370\uc774\ud130\uc14b \ub300\uc2dc\ubcf4\ub4dc ({len(all_rows)}\uac74)"
    ws_dashboard["B3"] = len(all_rows)
    ws_dashboard["B6"] = len(general_rows)
    ws_dashboard["B7"] = len(compliance_rows)
    ws_dashboard["B8"] = len(student_rows)

    dashboard_q_rows = [6, 7, 8, 9, 10]
    for row_num, question_type in zip(dashboard_q_rows, QUESTION_TYPES):
        ws_dashboard[f"D{row_num}"] = question_type
        ws_dashboard[f"E{row_num}"] = q_counter.get(question_type, 0)

    dashboard_d_rows = [12, 13, 14]
    for row_num, difficulty in zip(dashboard_d_rows, DIFFICULTIES):
        ws_dashboard[f"A{row_num}"] = difficulty
        ws_dashboard[f"B{row_num}"] = d_counter.get(difficulty, 0)

    ws_dashboard["D12"] = "\uc8fc\uc694 \uc548\ub0b4"
    ws_dashboard["E12"] = "\uc804 \uac74 \uae30\uc900\uc73c\ub85c metadata + hybrid chunk\ub97c \uc870\ud569\ud574 \uc790\ub3d9 \uc791\uc131\ud55c \ud559\uc2b5\ub370\uc774\ud130 \uc124\uacc4\uc11c\uc785\ub2c8\ub2e4."
    ws_dashboard["E13"] = "\ud398\ub974\uc18c\ub098, \uc9c8\ubb38\uc720\ud615, \ub09c\uc774\ub3c4\ub294 \uc758\uacb0\uc11c \uc81c\ubaa9\u00b7\uc704\ubc18\uc720\ud615\u00b7\uc870\uce58\uc720\ud615\u00b7chunk \uad6c\uc870\ub97c \uae30\ubc18\uc73c\ub85c \uaddc\uce59 \ubd84\ub958\ud588\uc2b5\ub2c8\ub2e4."

    wb.save(workbook_path)
    dump_workbook_json(wb, workbook_path.with_name("xlsx_dump.json"))

    print(f"updated_rows={len(all_rows)}")
    print(f"{PERSONA_GENERAL}={len(general_rows)}")
    print(f"{PERSONA_COMPLIANCE}={len(compliance_rows)}")
    print(f"{PERSONA_STUDENT}={len(student_rows)}")
    for question_type in QUESTION_TYPES:
        print(f"{question_type}={q_counter.get(question_type, 0)}")
    for difficulty in DIFFICULTIES:
        print(f"{difficulty}={d_counter.get(difficulty, 0)}")


if __name__ == "__main__":
    main()
